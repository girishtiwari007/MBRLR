"""Refresh Revenue Liability Portal static data from local IPAS/MB budget XLS files.

Default source folder:
  C:\\Users\\HP\\Downloads\\PORTAL DATA

This script is intentionally local. GitHub Pages can serve files, but it cannot
write parsed data back into this repository from the browser.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import xlrd
except ImportError as exc:  # pragma: no cover
    raise SystemExit("xlrd is required. Install it in Python before running this sync.") from exc

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


FY = "2026-2027"
FY_SHORT = "2026-27"
PORTAL_CODE_REVISION = "desktop-sync-readonly1"
IST = timezone(timedelta(hours=5, minutes=30))
MONTH_KEYS = ["apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan", "feb", "mar"]
MONTH_LABELS = ["APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC", "JAN", "FEB", "MAR"]
SKIP_DETAIL_PUS = {"72", "73", "74", "75", "98"}
DEMAND_BY_SMH = {
    "01": "03", "02": "04", "03": "05", "04": "06", "05": "07", "06": "08",
    "07": "09", "08": "10", "09": "11", "10": "12", "11": "13", "10N": "12N",
}
DEPT_BY_SMH = {
    "01": "PERSONNEL / STORE And Office Staff",
    "02": "ENGINEERING / PWAY",
    "03": "Mechanical LOCO Shed Roza",
    "04": "Electrical General / Mech C&W",
    "05": "S&T / TRD",
    "06": "MECHANICAL / Running Staff",
    "07": "OPERATING / Commercial",
    "08": "Operating Expenses - Fuel / Traction",
    "09": "MEDICAL",
    "10": "SECURITY",
    "11": "Pension and Retirement",
    "10N": "Suspense Heads",
}
SOURCE_NAMES = {
    "pu_budget": "PU-BUDGET.xls",
    "pu_month": "PU-MONTH-ACTUAL.xls",
    "detail_budget": "PU-DEPT-DEMAND-SMH-BUDGET.xls",
    "detail_actual": "PU-DEPT-DEMAND-SMH-ACTUAL.xls",
    "demand_budget": "DEMAND-SMH-BUGDET.xls",
    "demand_actual": "DEMAND-SMH-ACTUAL.xls",
}
ROLE_LABELS = {
    "pu_budget": "PU-wise Budget Available",
    "pu_month": "PU-wise Month-wise Actual",
    "detail_budget": "DEPT-Demand/SMH PU-wise Budget",
    "detail_actual": "DEPT-Demand/SMH PU-wise Month-wise Actual",
    "demand_budget": "Demand/SMH Budget Summary",
    "demand_actual": "Demand/SMH Month-wise Actual Summary",
}
TARGET_NAMES = {
    "pu_budget": "pu-budget.xls",
    "pu_month": "pu-month-actual.xls",
    "detail_budget": "pu-dept-demand-smh-budget.xls",
    "detail_actual": "pu-dept-demand-smh-actual.xls",
    "demand_budget": "demand-smh-budget.xls",
    "demand_actual": "demand-smh-actual.xls",
}


def norm_header(value) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def norm_code(value, width=2) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    if isinstance(value, float) and value.is_integer():
        raw = str(int(value))
    match = re.search(r"(\d+[A-Z]?)", raw.upper())
    if not match:
        return raw.upper()
    code = match.group(1)
    return code if code.endswith("N") else code.zfill(width)


def as_number(value) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "NIL"}:
        return 0
    neg = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    try:
        num = int(round(float(text)))
    except ValueError:
        return 0
    return -num if neg else num


class XlsxSheet:
    def __init__(self, path: Path):
        if openpyxl is None:
            raise SystemExit("openpyxl is required for .xlsx source files.")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        self.name = ws.title
        self.rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        self.nrows = len(self.rows)
        self.ncols = max((len(row) for row in self.rows), default=0)

    def cell_value(self, row: int, col_index: int):
        if row >= self.nrows or col_index is None or col_index >= len(self.rows[row]):
            return ""
        return self.rows[row][col_index]


def read_sheet(path: Path):
    if path.suffix.lower() == ".xlsx":
        return XlsxSheet(path)
    book = xlrd.open_workbook(str(path))
    return book.sheet_by_index(0)


def find_header(sheet, required):
    required = [norm_header(r) for r in required]
    for r in range(min(sheet.nrows, 20)):
        headers = [norm_header(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        if all(any(req in h for h in headers) for req in required):
            return r, headers
    raise ValueError(f"Header not found in {sheet.name}: {required}")


def col(headers, *patterns):
    for pattern in patterns:
        pattern = norm_header(pattern)
        for i, header in enumerate(headers):
            if pattern and pattern in header:
                return i
    return None


def month_col(headers, month_label):
    month_label = norm_header(month_label)
    for i, header in enumerate(headers):
        if month_label in header and "ACTUAL" in header:
            return i
    for i, header in enumerate(headers):
        if month_label in header:
            return i
    return None


def sheet_profile(path: Path):
    sh = read_sheet(path)
    best = {"path": path, "role": None, "headerRow": None, "headers": [], "score": 0, "reason": ""}
    for r in range(min(sh.nrows, 20)):
        headers = [norm_header(sh.cell_value(r, c)) for c in range(sh.ncols)]
        has_pu = any("PUCODE" in h for h in headers)
        has_dept = any("DEPARTMENTCODE" in h for h in headers)
        has_smh = any(h == "SMH" or h.endswith("SMH") or "SMH" in h for h in headers)
        has_au = any(h == "AU" for h in headers)
        has_bg = any("BGISL" in h or h in {"BUDGET", "OBA"} for h in headers)
        has_rg_stage = any(h in {"REA20262027", "RG20262027", "FME20262027", "FG20262027", "RG"} for h in headers)
        month_hits = sum(1 for label in MONTH_LABELS if any(label in h for h in headers))
        role = None
        score = 0
        reason = ""
        if has_dept and has_smh and has_pu and month_hits >= 3:
            role, score, reason = "detail_actual", 95, "Department + SMH + PU + month actual columns"
        elif has_dept and has_smh and has_pu and has_bg:
            role, score, reason = "detail_budget", 90, "Department + SMH + PU + BG/Budget columns"
        elif has_pu and not has_dept and not has_smh and month_hits >= 3:
            role, score, reason = "pu_month", 85, "PU + month actual columns"
        elif has_pu and not has_dept and has_bg:
            role, score, reason = "pu_budget", 80 + (5 if has_rg_stage else 0), "PU + BG/RG budget columns"
        elif has_au and has_smh and month_hits >= 3:
            role, score, reason = "demand_actual", 75, "AU + SMH + month actual columns"
        elif has_au and has_smh and has_bg:
            role, score, reason = "demand_budget", 70, "AU + SMH + BG/Budget columns"
        if score > best["score"]:
            best.update({"role": role, "headerRow": r, "headers": headers, "score": score, "reason": reason})
    return best


def discover_source_files(source_dir: Path):
    candidates = []
    for pattern in ("*.xls", "*.xlsx"):
        candidates.extend(p for p in source_dir.glob(pattern) if not p.name.startswith("~$"))
    if not candidates:
        raise SystemExit(f"No .xls/.xlsx files found in {source_dir}")

    profiles = []
    for path in candidates:
        try:
            profile = sheet_profile(path)
            if profile["role"]:
                profiles.append(profile)
        except Exception as exc:
            profiles.append({"path": path, "role": None, "score": 0, "reason": f"Not readable: {exc}"})

    selected = {}
    for role in SOURCE_NAMES:
        exact = source_dir / SOURCE_NAMES[role]
        if exact.exists():
            selected[role] = exact
            continue
        matches = [p for p in profiles if p.get("role") == role]
        if matches:
            matches.sort(key=lambda p: (p["score"], p["path"].stat().st_mtime), reverse=True)
            selected[role] = matches[0]["path"]

    missing = [ROLE_LABELS[r] for r in SOURCE_NAMES if r not in selected]
    if missing:
        detected = "\n".join(
            f"- {p['path'].name}: {ROLE_LABELS.get(p.get('role'), 'Not detected')} ({p.get('reason', '-')})"
            for p in profiles
        )
        raise SystemExit("Could not detect all required source files:\n"
                         + "\n".join(f"- {m}" for m in missing)
                         + "\n\nDetected files:\n" + detected)
    return selected, profiles


def discover_py_source_files(source_dir: Path):
    candidates = []
    for pattern in ("*.xls", "*.xlsx"):
        candidates.extend(p for p in source_dir.glob(pattern) if not p.name.startswith("~$"))
    profiles = []
    for path in candidates:
        try:
            profile = sheet_profile(path)
            if profile.get("role") in {"pu_budget", "pu_month"}:
                profiles.append(profile)
        except Exception:
            continue
    selected = {}
    for role in ("pu_budget", "pu_month"):
        matches = [p for p in profiles if p.get("role") == role]
        if matches:
            matches.sort(key=lambda p: (p["score"], p["path"].stat().st_mtime), reverse=True)
            selected[role] = matches[0]["path"]
    missing = [ROLE_LABELS[r] for r in ("pu_budget", "pu_month") if r not in selected]
    if missing:
        raise SystemExit("Could not detect Previous Year source files:\n" + "\n".join(f"- {m}" for m in missing))
    return selected, profiles


def load_existing_maps(root: Path):
    pu_names, dept_names = {}, {}
    detail_path = root / "assets/js/detail-data.js"
    if detail_path.exists():
        text = detail_path.read_text(encoding="utf-8").strip()
        try:
            data = json.loads(text[text.index("=") + 1:].rstrip(";").strip())
            for row in data.get("rows", []):
                pu_names.setdefault(str(row.get("puCode", "")).zfill(2), row.get("puName", ""))
                dept_names.setdefault(str(row.get("deptCode", "")).zfill(2), row.get("deptName", ""))
        except Exception:
            pass
    return pu_names, dept_names


def parse_pu_budget(path: Path):
    sh = read_sheet(path)
    hr, headers = find_header(sh, ["PUCODE"])
    c_pu = col(headers, "PUCODE")
    c_bg = col(headers, "BGISL20262027", "BGISL")
    c_rg = col(headers, "RG")
    c_actual = col(headers, "ACTUALSUPTO")
    out = {}
    for r in range(hr + 1, sh.nrows):
        code = norm_code(sh.cell_value(r, c_pu))
        if not code:
            continue
        out[code] = {
            "bg_isl": as_number(sh.cell_value(r, c_bg)) if c_bg is not None else 0,
            "rg": as_number(sh.cell_value(r, c_rg)) if c_rg is not None else 0,
            "actuals_till": as_number(sh.cell_value(r, c_actual)) if c_actual is not None else 0,
        }
    detail_rows = [values for code, values in out.items() if code != "TOTAL"]
    out["TOTAL"] = {
        field: sum(values[field] for values in detail_rows)
        for field in ("bg_isl", "rg", "actuals_till")
    }
    return out


def parse_pu_month(path: Path, start_year=2026):
    sh = read_sheet(path)
    hr, headers = find_header(sh, ["PUCODE"])
    c_pu = col(headers, "PUCODE")
    month_cols = [month_col(headers, f"{m} {start_year}" if i <= 8 else f"{m} {start_year + 1}") for i, m in enumerate(MONTH_LABELS)]
    out = {}
    latest = -1
    for r in range(hr + 1, sh.nrows):
        code = norm_code(sh.cell_value(r, c_pu))
        if not code:
            continue
        vals = {}
        for i, c in enumerate(month_cols):
            vals[MONTH_KEYS[i]] = as_number(sh.cell_value(r, c)) if c is not None else 0
            if vals[MONTH_KEYS[i]]:
                latest = max(latest, i)
        out[code] = vals
    detail_rows = [values for code, values in out.items() if code != "TOTAL"]
    out["TOTAL"] = {
        month: sum(values[month] for values in detail_rows)
        for month in MONTH_KEYS
    }
    return out, latest


def parse_detail(budget_path: Path, actual_path: Path, pu_names, dept_names):
    def key_from(sh, headers, r):
        c_dept, c_smh, c_pu = col(headers, "DEPARTMENTCODE"), col(headers, "SMH"), col(headers, "PUCODE")
        return norm_code(sh.cell_value(r, c_dept)), norm_code(sh.cell_value(r, c_smh)), norm_code(sh.cell_value(r, c_pu))

    bsh = read_sheet(budget_path)
    bhr, bheaders = find_header(bsh, ["DEPARTMENTCODE", "SMH", "PUCODE"])
    c_bg = col(bheaders, "BGISL", "BUDGET")
    budget = {}
    for r in range(bhr + 1, bsh.nrows):
        dept, smh, pu = key_from(bsh, bheaders, r)
        if not dept or not smh or not pu or dept == "00" or pu in SKIP_DETAIL_PUS:
            continue
        budget[(dept, smh, pu)] = budget.get((dept, smh, pu), 0) + (as_number(bsh.cell_value(r, c_bg)) if c_bg is not None else 0)

    ash = read_sheet(actual_path)
    ahr, aheaders = find_header(ash, ["DEPARTMENTCODE", "SMH", "PUCODE"])
    month_cols = [month_col(aheaders, f"{m} 2026" if i <= 8 else f"{m} 2027") for i, m in enumerate(MONTH_LABELS)]
    actual = {}
    for r in range(ahr + 1, ash.nrows):
        dept, smh, pu = key_from(ash, aheaders, r)
        if not dept or not smh or not pu or dept == "00" or pu in SKIP_DETAIL_PUS:
            continue
        vals = actual.setdefault((dept, smh, pu), {m: 0 for m in MONTH_KEYS})
        for i, c in enumerate(month_cols):
            if c is not None:
                vals[MONTH_KEYS[i]] += as_number(ash.cell_value(r, c))

    rows = []
    for dept, smh, pu in sorted(set(budget) | set(actual)):
        months = actual.get((dept, smh, pu), {m: 0 for m in MONTH_KEYS})
        rows.append({
            "deptCode": dept,
            "deptName": dept_names.get(dept, f"DEPARTMENT {dept}"),
            "smh": f"SMH - {smh}",
            "puCode": pu,
            "puName": pu_names.get(pu, f"PU {pu}"),
            "budget": budget.get((dept, smh, pu), 0),
            "actualTill": sum(months.values()),
            "months": months,
        })
    totals = {
        "budget": sum(r["budget"] for r in rows),
        "actualTill": sum(r["actualTill"] for r in rows),
        "months": {m: sum(r["months"][m] for r in rows) for m in MONTH_KEYS},
    }
    return {"source": "PORTAL DATA local sync", "rules": "Department 00 and PU 72/73/74/75/98 skipped for display", "monthLabels": MONTH_LABELS, "monthKeys": MONTH_KEYS, "totals": totals, "rows": rows}


def fy_month_label(idx: int) -> str:
    if idx < 0:
        return "No completed month"
    return f"{MONTH_LABELS[idx]} {'2026' if idx <= 8 else '2027'}"


def current_fy_month_idx(now: datetime) -> int:
    month = now.month
    return month - 4 if month >= 4 else month + 8


def bp_mode_note(completed: int, current_idx: int, latest_idx: int) -> str:
    completed_label = fy_month_label(completed - 1)
    current_label = fy_month_label(current_idx)
    latest_label = fy_month_label(latest_idx)
    return f"Completed through {completed_label}; {current_label} is current running month; latest uploaded actual month detected as {latest_label}"


def parse_demand(budget_path: Path, actual_path: Path, generated_at: str, completed: int, current_idx: int, latest_idx: int):
    bsh = read_sheet(budget_path)
    bhr, bheaders = find_header(bsh, ["AU", "SMH"])
    c_smh, c_bg = col(bheaders, "SMH"), col(bheaders, "BGISL", "BUDGET", "OBA")
    oba = {}
    for r in range(bhr + 1, bsh.nrows):
        smh = norm_code(bsh.cell_value(r, c_smh))
        if smh and smh not in {"TOTAL", "GRANDTOTAL"}:
            oba[smh] = oba.get(smh, 0) + (as_number(bsh.cell_value(r, c_bg)) if c_bg is not None else 0)

    ash = read_sheet(actual_path)
    ahr, aheaders = find_header(ash, ["AU", "SMH"])
    c_smh = col(aheaders, "SMH")
    month_cols = [month_col(aheaders, f"{m} 2026" if i <= 8 else f"{m} 2027") for i, m in enumerate(MONTH_LABELS)]
    actual = {}
    for r in range(ahr + 1, ash.nrows):
        smh = norm_code(ash.cell_value(r, c_smh))
        if not smh or smh in {"TOTAL", "GRANDTOTAL"}:
            continue
        vals = actual.setdefault(smh, {m: 0 for m in MONTH_KEYS})
        for i, c in enumerate(month_cols):
            if c is not None:
                vals[MONTH_KEYS[i]] += as_number(ash.cell_value(r, c))

    rows = []
    for smh in sorted(set(oba) | set(actual), key=lambda x: (x.endswith("N"), x)):
        months = actual.get(smh, {m: 0 for m in MONTH_KEYS})
        row_oba = oba.get(smh, 0)
        ae = sum(months[m] for m in MONTH_KEYS[:completed])
        bp = round(row_oba / 12 * completed)
        rows.append({
            "demand": DEMAND_BY_SMH.get(smh, smh),
            "smh": smh,
            "dept": DEPT_BY_SMH.get(smh, f"SMH {smh}"),
            "description": DEPT_BY_SMH.get(smh, f"SMH {smh}"),
            "oba": row_oba,
            "ae": ae,
            "months": months,
            "bp": bp,
            "variation": ae - bp,
            "bpPct": (ae / bp * 100) if bp else (100 if ae == 0 else 999),
            "budgetRemaining": row_oba - ae,
            "obaUtil": (ae / row_oba * 100) if row_oba else (100 if ae == 0 else 999),
        })
    main = [r for r in rows if not r["demand"].endswith("N") and not r["smh"].endswith("N")]
    totals = {k: sum(r[k] for r in main) for k in ["oba", "bp", "ae", "variation", "budgetRemaining"]}
    totals["bpPct"] = totals["ae"] / totals["bp"] * 100 if totals["bp"] else 0
    totals["obaUtil"] = totals["ae"] / totals["oba"] * 100 if totals["oba"] else 0
    return {
        "fy": FY,
        "sourceBudget": SOURCE_NAMES["demand_budget"],
        "sourceActual": SOURCE_NAMES["demand_actual"],
        "sourceCode": "Built-in Demand/SMH mapping",
        "generatedAt": generated_at,
        "asOn": fy_month_label(completed - 1),
        "completedMonths": completed,
        "note": f"OBA uses uploaded BG_ISL/OBA. BP and default Demand/SMH summary use {completed} completed month(s) through {fy_month_label(completed - 1)}. {fy_month_label(current_idx)} is treated as current running month. Latest uploaded actual month detected as {fy_month_label(latest_idx)}. Demand 12N/10N Suspense Heads is shown separately and is not netted from main total.",
        "rows": rows,
        "totals": totals,
    }


def replace_js_assignment(text: str, name: str, value: str) -> str:
    pattern = re.compile(rf"let {name} = .*?;\n", re.S)
    return pattern.sub(f"let {name} = {value};\n", text, count=1)


def validate_generated_outputs(root: Path):
    app = (root / "assets/js/app.js").read_text(encoding="utf-8")
    def assignment(name):
        match = re.search(rf"let {name} = (.*?);\n", app, re.S)
        if not match:
            raise RuntimeError(f"Generated portal assignment missing: {name}")
        return json.loads(match.group(1))

    budget = assignment("BUDGET")
    month = assignment("MONTH")
    mismatches = []
    for code, values in month.items():
        if code == "TOTAL" or code not in budget:
            continue
        month_sum = sum(float(values.get(key, 0) or 0) for key in MONTH_KEYS)
        if abs(month_sum - float(budget[code].get("actuals_till", 0) or 0)) > 0.5:
            mismatches.append(code)
    if mismatches:
        raise RuntimeError("PU actual/month mismatch after generation: " + ", ".join(mismatches[:15]))

    def window_data(path):
        text = path.read_text(encoding="utf-8").strip()
        return json.loads(text[text.index("=") + 1:].rstrip(";").strip())

    detail = window_data(root / "assets/js/detail-data.js")
    detail_month_sum = sum(float(detail.get("totals", {}).get("months", {}).get(key, 0) or 0) for key in MONTH_KEYS)
    if abs(detail_month_sum - float(detail.get("totals", {}).get("actualTill", 0) or 0)) > 0.5:
        raise RuntimeError("DEPT/Demand detail actual total does not equal its month total")
    demand = window_data(root / "assets/js/demand-smh-data.js")
    operational = [r for r in demand.get("rows", []) if str(r.get("smh", "")).upper() != "10N"]
    demand_ae = sum(float(r.get("ae", 0) or 0) for r in operational)
    if abs(demand_ae - float(demand.get("totals", {}).get("ae", 0) or 0)) > 0.5:
        raise RuntimeError("Demand/SMH main AE total does not reconcile")
    return {
        "ok": True,
        "puMonthMismatches": 0,
        "detailRows": len(detail.get("rows", [])),
        "detailActual": detail.get("totals", {}).get("actualTill", 0),
        "demandRows": len(demand.get("rows", [])),
        "demandMainAE": demand.get("totals", {}).get("ae", 0),
    }


def write_outputs(root: Path, source_dir: Path, github_dir: Path | None, py_source_dir: Path | None = None):
    now = datetime.now(IST).replace(microsecond=0)
    generated_at = now.isoformat()
    source_paths, detected_profiles = discover_source_files(source_dir)
    py_source_paths, py_profiles = ({}, [])
    if py_source_dir:
        py_source_paths, py_profiles = discover_py_source_files(py_source_dir)
    source_hasher = hashlib.sha256()
    all_revision_sources = [(f"cy:{role}", path) for role, path in source_paths.items()]
    all_revision_sources += [(f"py:{role}", path) for role, path in py_source_paths.items()]
    for role, path in sorted(all_revision_sources):
        source_hasher.update(role.encode("utf-8"))
        with path.open("rb") as source_file:
            for chunk in iter(lambda: source_file.read(1024 * 1024), b""):
                source_hasher.update(chunk)
    source_revision = source_hasher.hexdigest()[:12]
    version = f"{now.strftime('%Y%m%d')}-{PORTAL_CODE_REVISION}-autoexports-{source_revision}"

    pu_names, dept_names = load_existing_maps(root)
    budget = parse_pu_budget(source_paths["pu_budget"])
    month, latest_idx = parse_pu_month(source_paths["pu_month"])
    for code, vals in month.items():
        budget.setdefault(code, {"bg_isl": 0, "rg": 0, "actuals_till": 0})
        budget[code]["actuals_till"] = sum(vals.values())
    detail = parse_detail(source_paths["detail_budget"], source_paths["detail_actual"], pu_names, dept_names)
    budget_py = month_py = None
    if py_source_paths:
        budget_py = parse_pu_budget(py_source_paths["pu_budget"])
        month_py, _ = parse_pu_month(py_source_paths["pu_month"], start_year=2025)
        for code, vals in month_py.items():
            if code == "TOTAL":
                continue
            budget_py.setdefault(code, {"bg_isl": 0, "rg": 0, "actuals_till": 0})
            budget_py[code]["actuals_till"] = sum(vals.values())
    current_idx = current_fy_month_idx(now)
    completed_count = max(0, min(latest_idx, current_idx - 1) + 1)
    mode_note = bp_mode_note(completed_count, current_idx, latest_idx)
    demand = parse_demand(source_paths["demand_budget"], source_paths["demand_actual"], generated_at, completed_count, current_idx, latest_idx)
    detail["generatedAt"] = generated_at

    (root / "assets/js/detail-data.js").write_text("window.DETAIL_SMH_DATA = " + json.dumps(detail, separators=(",", ":")) + ";\n", encoding="utf-8", newline="\n")
    (root / "assets/js/demand-smh-data.js").write_text("window.DEMAND_SMH_SUMMARY_DATA = " + json.dumps(demand, separators=(",", ":")) + ";\n", encoding="utf-8", newline="\n")

    app_path = root / "assets/js/app.js"
    app = app_path.read_text(encoding="utf-8")
    app = re.sub(r"const ASSET_VERSION = '[^']+';", f"const ASSET_VERSION = '{version}';", app, count=1)
    app = replace_js_assignment(app, "BUDGET", json.dumps(budget, separators=(",", ":")))
    app = replace_js_assignment(app, "MONTH", json.dumps(month, separators=(",", ":")))
    if budget_py is not None and month_py is not None:
        app = replace_js_assignment(app, "BUDGET_PY", json.dumps(budget_py, separators=(",", ":")))
        app = replace_js_assignment(app, "MONTH_PY", json.dumps(month_py, separators=(",", ":")))
    stamp = now.strftime("%d-%b-%Y")
    latest_label = fy_month_label(latest_idx)
    latest_month_short = MONTH_LABELS[latest_idx] if latest_idx >= 0 else "latest"
    source_updates = {
        "budgetCY": f"Repository source refreshed from PORTAL DATA on {stamp}; actual till date aligned to APR-{latest_month_short} month-wise file.",
        "monthCY": f"Repository source refreshed from PORTAL DATA on {stamp}; latest loaded month {latest_label}.",
        "smhBudgetCY": f"Repository source refreshed from PORTAL DATA on {stamp}.",
        "smhMonthCY": f"Repository source refreshed from PORTAL DATA on {stamp}; latest loaded month {latest_label}.",
        "demandSmhCY": f"Repository source refreshed from PORTAL DATA on {stamp}. {mode_note}. Demand 12N/10N Suspense Heads is shown separately.",
    }
    for key, remark in source_updates.items():
        app = re.sub(
            rf"({key}: \{{label:'[^']+', fy:'[^']+', source:'[^']+', used:'[^']+', remarks:')[^']*('}})",
            rf"\1{remark}\2",
            app,
            count=1,
        )
    if py_source_paths:
        py_stamp = now.strftime("%d-%b-%Y")
        for key in ("budgetPY", "monthPY"):
            app = re.sub(
                rf"({key}: \{{label:'[^']+', fy:'[^']+', source:'[^']+', used:'[^']+', remarks:')[^']*('\}})",
                rf"\1Previous year repository source refreshed by MBRLR Local Sync on {py_stamp}.\2",
                app,
                count=1,
            )
    if "const FY_MONTHS =" not in app:
        app = app.replace("const DEFAULT_DATA_AS_ON_DATE", "const FY_MONTHS = ['apr','may','jun','jul','aug','sep','oct','nov','dec','jan','feb','mar'];\nconst FY_MONTH_LABELS = ['APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC','JAN','FEB','MAR'];\nconst DEFAULT_DATA_AS_ON_DATE", 1)
    app = re.sub(r"const DEFAULT_DATA_AS_ON_DATE = new Date\('[^']+'\);", f"const DEFAULT_DATA_AS_ON_DATE = new Date('{generated_at}');", app, count=1)
    app = re.sub(r"const RLP_BUILD_ID = '[^']+';", f"const RLP_BUILD_ID = 'rlp-mbd-{generated_at[:10]}-{PORTAL_CODE_REVISION}-{source_revision}';", app, count=1)
    app_path.write_text(app, encoding="utf-8", newline="\n")

    index_path = root / "index.html"
    index = index_path.read_text(encoding="utf-8")
    fallback = datetime.fromisoformat(generated_at).strftime("%d-%b-%Y %H:%M")
    index = re.sub(r'<meta name="portal-build" content="[^"]+">', f'<meta name="portal-build" content="rlp-mbd-{generated_at[:10]}-{PORTAL_CODE_REVISION}-{source_revision}">', index, count=1)
    index = re.sub(r'As on: [0-9]{2}-[A-Za-z]{3}-[0-9]{4} [0-9]{2}:[0-9]{2}', f"As on: {fallback}", index, count=1)
    index = re.sub(r'v=2026[0-9A-Za-z-]+', f"v={version}", index)
    index_path.write_text(index, encoding="utf-8", newline="\n")

    target_source = root / "data/mb-budget-sync/source-files/2026-2027"
    target_source.mkdir(parents=True, exist_ok=True)
    source_file_entries = []
    for key, src in source_paths.items():
        dst = target_source / TARGET_NAMES[key]
        if src.resolve() != dst.resolve():
            shutil.copy2(src, dst)
        stat = src.stat()
        source_file_entries.append({
            "name": TARGET_NAMES[key],
            "originalName": src.name,
            "role": key,
            "roleLabel": ROLE_LABELS[key],
            "relativePath": f"data/source-files/2026-2027/{TARGET_NAMES[key]}",
            "size": stat.st_size,
            "modifiedAt": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
            "sha256": hashlib.sha256(src.read_bytes()).hexdigest(),
            "targetPath": f"data/mb-budget-sync/source-files/2026-2027/{TARGET_NAMES[key]}",
        })
    py_source_file_entries = []
    if py_source_paths:
        py_target = root / "data/mb-budget-sync/source-files/2025-2026"
        py_target.mkdir(parents=True, exist_ok=True)
        for key, src in py_source_paths.items():
            target_name = TARGET_NAMES[key]
            dst = py_target / target_name
            if src.resolve() != dst.resolve():
                shutil.copy2(src, dst)
            stat = src.stat()
            py_source_file_entries.append({
                "name": target_name,
                "originalName": src.name,
                "role": f"py_{key}",
                "roleLabel": f"Previous Year {ROLE_LABELS[key]}",
                "size": stat.st_size,
                "modifiedAt": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                "sha256": hashlib.sha256(src.read_bytes()).hexdigest(),
                "targetPath": f"data/mb-budget-sync/source-files/2025-2026/{target_name}",
            })

    processed = root / "data/mb-budget-sync/processed"
    processed.mkdir(parents=True, exist_ok=True)
    current_payload = {
        "demand": {"title": "Demand / SMH Wise Current Year", "rows": [
            {"Name": f"Demand {r['demand']} / {r['smh']}", "OBA": r["oba"], "BP": r["bp"], "AE": r["ae"], "Variation": r["variation"], "BPPercent": r["bpPct"], "Remaining": r["budgetRemaining"], "OBAPercent": r["obaUtil"], "Months": demand["completedMonths"], "Department": r["dept"]}
            for r in demand["rows"]
        ] + [{"Name": "Total", "OBA": demand["totals"]["oba"], "BP": demand["totals"]["bp"], "AE": demand["totals"]["ae"], "Variation": demand["totals"]["variation"], "BPPercent": demand["totals"]["bpPct"], "Remaining": demand["totals"]["budgetRemaining"], "OBAPercent": demand["totals"]["obaUtil"], "Months": demand["completedMonths"]}]},
        "generatedAt": generated_at,
        "source": "PORTAL DATA local sync",
    }
    (processed / "current_payload.js").write_text("window.CURRENT_PAYLOAD = " + json.dumps(current_payload, separators=(",", ":")) + ";\n", encoding="utf-8", newline="\n")
    reports = {"budget": {"demand": {f"Demand {r['demand']} / SMH {r['smh']}": {FY_SHORT: {"oba": r["oba"], "ae": r["ae"], "bp": r["bp"]}} for r in demand["rows"]}}, "summary": {"generatedAt": generated_at, "latestMonth": fy_month_label(latest_idx), "bpMode": mode_note, "demandRows": len(demand["rows"]), "demandMainOBA": demand["totals"]["oba"], "demandMainBP": demand["totals"]["bp"], "demandMainAE": demand["totals"]["ae"]}}
    (processed / "reports-data.json").write_text(json.dumps(reports, indent=2), encoding="utf-8", newline="\n")
    detected_map = {role: {"detectedFile": source_paths[role].name, "storedAs": TARGET_NAMES[role], "label": ROLE_LABELS[role]} for role in SOURCE_NAMES}
    (processed / "year-sources.json").write_text(json.dumps({"financialYear": FY, "generatedAt": generated_at, "sourceFolder": str(source_dir), "files": detected_map}, indent=2), encoding="utf-8", newline="\n")

    summary = {
        "puBudgetRows": len(budget),
        "puMonthRows": len(month),
        "detailRows": len(detail["rows"]),
        "detailBudget": detail["totals"]["budget"],
        "detailActual": detail["totals"]["actualTill"],
        "demandRows": len(demand["rows"]),
        "latestMonth": reports["summary"]["latestMonth"],
        "demandMainAECompleted": demand["totals"]["ae"],
        "demandMainBPCompleted": demand["totals"]["bp"],
        "pyBudgetRows": len(budget_py) if budget_py is not None else 0,
        "pyMonthRows": len(month_py) if month_py is not None else 0,
    }
    calculation_validation = validate_generated_outputs(root)
    manifest = {
        "ok": True,
        "mode": "local-portal-sync",
        "sourceRepo": str(source_dir),
        "targetRepo": str(root),
        "sourceRevision": source_revision,
        "assetVersion": version,
        "calculationValidation": calculation_validation,
        "financialYear": FY,
        "generatedAt": generated_at,
        "syncedAt": generated_at,
        "confirmedAt": generated_at,
        "counts": {"sourceFiles": 6 + len(py_source_file_entries), "processedFiles": 3, "frFiles": 0, "totalFiles": 11 + len(py_source_file_entries)},
        "sourceFiles": source_file_entries + py_source_file_entries,
        "detectedProfiles": [
            {
                "file": p["path"].name,
                "role": p.get("role"),
                "roleLabel": ROLE_LABELS.get(p.get("role"), "Not detected"),
                "score": p.get("score", 0),
                "reason": p.get("reason", ""),
            }
            for p in detected_profiles
        ],
        "processedFiles": [
            {"name": "current_payload.js", "targetPath": "data/mb-budget-sync/processed/current_payload.js"},
            {"name": "reports-data.json", "targetPath": "data/mb-budget-sync/processed/reports-data.json"},
            {"name": "year-sources.json", "targetPath": "data/mb-budget-sync/processed/year-sources.json"},
        ],
        "selectedTargets": [f["targetPath"] for f in source_file_entries + py_source_file_entries],
        "actionLog": [{"at": generated_at, "action": "Local source files detected by sheet contents, parsed, verified and written into static portal data", "status": "confirmed", "files": [f"{f['roleLabel']}: {f['originalName']}" for f in source_file_entries], "summary": summary}],
        "bpMode": mode_note,
    }
    (root / "data/mb-budget-sync/sync-manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8", newline="\n")
    (root / "data/mb-budget-sync/sync-log.json").write_text(json.dumps(manifest["actionLog"], indent=2), encoding="utf-8", newline="\n")

    if github_dir and github_dir.resolve() != root.resolve():
        for rel in [
            "index.html",
            "assets/css/main.css",
            "assets/js/app.js",
            "assets/js/detail-data.js",
            "assets/js/demand-smh-data.js",
            "data/mb-budget-sync",
            "tools/local_portal_sync.py",
            "tools/mbrlr_sync_gui.py",
            "tools/local_portal_server.py",
            "START-LOCAL-PORTAL-SERVER.bat",
            "START-MBRLR-LOCAL-SYNC.bat",
            "RUN-LOCAL-DATA-SYNC.bat",
            "README.md",
        ]:
            src = root / rel
            dst = github_dir / rel
            if src.is_dir():
                shutil.copytree(src, dst, dirs_exist_ok=True)
            else:
                dst.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(src, dst)

    assert detail["totals"]["actualTill"] == sum(detail["totals"]["months"].values())
    assert len(source_file_entries) == 6
    if py_source_paths:
        assert len(py_source_file_entries) == 2
        assert all(abs((budget_py.get(code) or {}).get("actuals_till", 0) - sum(values.values())) < 0.5 for code, values in month_py.items() if code != "TOTAL")
    return summary


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=r"C:\Users\HP\Downloads\PORTAL DATA")
    parser.add_argument("--root", default=str(Path(__file__).resolve().parents[1]))
    parser.add_argument("--github", default="")
    parser.add_argument("--py-source", default="", help="Optional Previous Year folder containing PU budget and PU month-wise actual files")
    args = parser.parse_args()
    summary = write_outputs(
        Path(args.root),
        Path(args.source),
        Path(args.github) if args.github else None,
        Path(args.py_source) if args.py_source else None,
    )
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
