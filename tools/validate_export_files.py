"""Validate portal-generated XLSX, PPTX and PDF fixed layout rules."""
from pathlib import Path
import sys
import xml.etree.ElementTree as ET
import zipfile

import fitz
from openpyxl import load_workbook


ROOT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent.parent / ".export-validation"
files = {suffix: next(ROOT.glob(f"*{suffix}")) for suffix in (".xlsx", ".pptx", ".pdf")}

# Excel: every explicit font is >=10 pt and every worksheet prints landscape, one page wide.
book = load_workbook(files[".xlsx"], read_only=False, data_only=False)
for sheet in book.worksheets:
    assert sheet.page_setup.orientation == "landscape", f"{sheet.title}: not landscape"
    assert sheet.page_setup.fitToWidth == 1, f"{sheet.title}: not fit to one page wide"
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and cell.font.sz is not None:
                assert float(cell.font.sz) >= 10, f"{sheet.title}!{cell.coordinate}: {cell.font.sz} pt"

# PowerPoint: package XML is valid, text >=10 pt, and all shapes remain inside 16:9 canvas.
with zipfile.ZipFile(files[".pptx"]) as package:
    names = package.namelist()
    for name in names:
        if name.endswith((".xml", ".rels")):
            ET.fromstring(package.read(name))
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    for name in (n for n in names if n.startswith("ppt/slides/slide") and n.endswith(".xml")):
        root = ET.fromstring(package.read(name))
        for props in root.findall(".//a:rPr", ns):
            assert int(props.attrib.get("sz", "1000")) >= 1000, f"{name}: text below 10 pt"
        for xfrm in root.findall(".//a:xfrm", ns):
            off, ext = xfrm.find("a:off", ns), xfrm.find("a:ext", ns)
            if off is not None and ext is not None:
                assert int(off.attrib["x"]) + int(ext.attrib["cx"]) <= 12192000, f"{name}: shape exceeds width"
                assert int(off.attrib["y"]) + int(ext.attrib["cy"]) <= 6858000, f"{name}: shape exceeds height"

# PDF: all pages are landscape and extracted text spans use at least 10 pt.
pdf = fitz.open(files[".pdf"])
for page_number, page in enumerate(pdf, 1):
    assert page.rect.width > page.rect.height, f"PDF page {page_number}: not landscape"
    for block in page.get_text("dict")["blocks"]:
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                if span.get("text", "").strip():
                    assert float(span["size"]) >= 9.95, f"PDF page {page_number}: {span['size']} pt"

print(f"PASS: {len(book.worksheets)} Excel sheets, {len(pdf)} PDF pages, and PowerPoint package validated")
